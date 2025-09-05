import tkinter as tk
from tkinter import ttk 
from PIL import Image, ImageTk
import os
import cv2
import threading
import openpyxl
from pygrabber.dshow_graph import FilterGraph
import pytesseract
import re
import numpy as np
import pyodbc
import sys
import time

pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

class App(tk.Tk):
        def __init__(self):
            super().__init__()
            self.title("Warenein- und -ausgang")
            
            # Bildschirmgröße holen
            screen_width = self.winfo_screenwidth()
            screen_height = self.winfo_screenheight()
            self.geometry(f"{screen_width}x{screen_height//2}+0+0")
            
            # Maße des Logos
            breite = 120
            hoehe = 60

            # Logo laden und scalieren
            self.load_and_scale_logo(breite, hoehe)

            # Maße der Bilder (Piktogramm)
            breite = 450
            hoehe = 450
            
            # Bilder laden und skalieren
            self.load_and_scale_images(breite, hoehe)
            
            # Frames (Seiten)
            self.startseite = tk.Frame(self)
            self.wareneingang_seite = tk.Frame(self)
            self.warenausgang_seite = tk.Frame(self)

             # Webcam suchen und initialisieren (in separatem Thread)
            self.cap = None
            threading.Thread(target=self.initialize_webcam, daemon=True).start()

            # OCR-States (nur aktiv auf Wareneingang/-ausgang)
            self.last_frame = None
            self.ocr_results = []
            self.stop_event = None

            # Video-/OCR-Settings (max. 720p, OCR bis 960px Breite)
            self.capture_candidates = [(1280, 720), (640, 480)]
            self.capture_fps = 30
            self.ocr_max_width = 960

            # DB-Pfad und Cache (für Karton/Beutel)
            self.db_path = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "datenbank", "Datenbank1.accdb"))
            self.db_cache_karton_beutel = {}  # key: normierte Artikelnummer -> (karton, beutel)

            # Diagnose: aktiver Interpreter und DB
            print(f"[DB] Python-Exe: {sys.executable}")
            print(f"[DB] DB-Pfad: {self.db_path} (exists={os.path.exists(self.db_path)})")
            try:
                print(f"[DB] ODBC-Treiber: { [d for d in pyodbc.drivers() if 'Access Driver' in d] }")
            except Exception as e:
                print(f"[DB] pyodbc/ODBC-Treiber nicht abfragbar: {e}")
            
            # Artikel-Daten
            self.artikel_dict_eingang = []           # Excel-Daten Wareneingang
            self.artikel_dict_ausgang = []           # Excel-Daten Warenausgang
            self.detected_articles_eingang = []      # (bleibt, falls später genutzt)
            self.detected_articles_ausgang = []      # (bleibt, falls später genutzt)
            self.detected_set_eingang = set()       # verhindert doppelte Einträge in der Wareneingang-Tabelle
            self.detected_set_ausgang = set()       # verhindert doppelte Einträge in der Warenausgang-Tabelle

            # Seiten aufbauen
            self.build_startseite()
            self.build_wareneingang()
            self.build_warenausgang()
            self.show_startseite()

            # Debugging-Hotkey (F9)
            self.bind("<F9>", lambda e: self.debug_print_article_dicts())

        #------------------------------------------------------- GUI ---------------------------------------------------------

        def load_and_scale_images(self, breite, hoehe):
            """Lädt und skaliert die Wareneingang- und Warenausgang-Bilder"""
            try:
                eingang_img = Image.open("../bilder/wareneingang.png").resize((breite, hoehe), Image.Resampling.LANCZOS)
                ausgang_img = Image.open("../bilder/warenausgang.png").resize((breite, hoehe), Image.Resampling.LANCZOS)
                self.eingang_photo = ImageTk.PhotoImage(eingang_img)
                self.ausgang_photo = ImageTk.PhotoImage(ausgang_img)
            except FileNotFoundError as e:
                print(f"Bilddatei nicht gefunden: {e}")
            except Exception as e:
                print(f"Fehler beim Laden der Bilder: {e}")

        def load_and_scale_logo(self, breite, hoehe):
            """Lädt und skaliert das Logo"""
            try:
                logo_img = Image.open("../bilder/logo.png").resize((breite, hoehe), Image.Resampling.LANCZOS)
                self.logo_photo = ImageTk.PhotoImage(logo_img)
            except FileNotFoundError as e:
                print(f"Logo-Datei nicht gefunden: {e}")
            except Exception as e:
                print(f"Fehler beim Laden des Logos: {e}")

        def build_startseite(self):
            btn_eingang = tk.Button(self.startseite, image=self.eingang_photo,
                                    command=self.show_wareneingang, borderwidth=0)
            btn_ausgang = tk.Button(self.startseite, image=self.ausgang_photo,
                                    command=self.show_warenausgang, borderwidth=0)
            btn_eingang.pack(side="left", expand=True, fill="both")
            btn_ausgang.pack(side="right", expand=True, fill="both")
            self.add_logo(self.startseite)

        def show_startseite(self):
            self.startseite.pack(expand=True, fill="both")
            self.wareneingang_seite.pack_forget()
            self.warenausgang_seite.pack_forget()
            self.current_page = "startseite"
            # OCR und Webcam auf Startseite stoppen
            self.stop_ocr()
            self.stop_webcam_stream()

        def add_logo(self, parent_frame):
            logo_label = tk.Label(parent_frame, image=self.logo_photo, borderwidth=0)
            logo_label.image = self.logo_photo
            logo_label.pack(side="right", anchor="se", padx=10, pady=10)

        def build_wareneingang(self):
            tk.Label(self.wareneingang_seite, text="Wareneingang", font=("Arial", 30)).pack(pady=20)
            main_frame = tk.Frame(self.wareneingang_seite)
            main_frame.pack(expand=True, fill="both")

            left_frame = tk.Frame(main_frame)
            left_frame.pack(side="left", expand=True, fill="both", padx=20, pady=20)

            self.dropdown_var_eingang = tk.StringVar()
            self.dropdown_eingang = ttk.Combobox(left_frame, textvariable=self.dropdown_var_eingang, width=30)
            excel_files_eingang = self.load_excel_files("../eingang")
            self.dropdown_eingang['values'] = excel_files_eingang
            if excel_files_eingang:
                self.dropdown_eingang.set(excel_files_eingang[0])
                filepath = os.path.join("../eingang", excel_files_eingang[0])
                self.load_excel_data(filepath, "eingang")
            self.dropdown_eingang.bind('<Button-1>', self.refresh_dropdown_eingang)
            self.dropdown_eingang.bind('<<ComboboxSelected>>', self.on_excel_select_eingang)
            self.dropdown_eingang.pack(pady=10)

            tk.Label(left_frame, text="Erfasste Artikel", font=("Arial", 14)).pack(pady=(20,5))
            columns = ("artikelnummer", "menge", "karton", "beutel", "status")
            self.tree_eingang = ttk.Treeview(left_frame, columns=columns, show="headings", height=15)
            self.tree_eingang.heading("artikelnummer", text="Artikelnummer")
            self.tree_eingang.heading("menge", text="Menge")
            self.tree_eingang.heading("karton", text="Karton")
            self.tree_eingang.heading("beutel", text="Beutel")
            self.tree_eingang.heading("status", text="Status")
            self.tree_eingang.column("artikelnummer", width=120, anchor="center")
            self.tree_eingang.column("menge", width=80, anchor="center")
            self.tree_eingang.column("karton", width=80, anchor="center")
            self.tree_eingang.column("beutel", width=80, anchor="center")
            self.tree_eingang.column("status", width=80, anchor="center")
            self.tree_eingang.pack(expand=True, fill="both")

            self.right_frame_eingang = tk.Frame(main_frame, bg="black")
            self.right_frame_eingang.pack(side="right", expand=False, fill="y", padx=20, pady=20)

            button_frame = tk.Frame(self.wareneingang_seite)
            button_frame.pack(pady=10)
            tk.Button(button_frame, text="Drucken", command=self.on_drucken_eingang).pack(side="left", padx=5)  # geändert
            tk.Button(button_frame, text="Zurück", command=self.show_startseite).pack(side="left", padx=5)
            self.add_logo(self.wareneingang_seite)

        def show_wareneingang(self):
            self.current_page = "eingang"
            self.startseite.pack_forget()
            self.warenausgang_seite.pack_forget()
            self.wareneingang_seite.pack(expand=True, fill="both")
            self.check_webcam_for_page()
            self.start_webcam_stream(self.right_frame_eingang)
            self.start_ocr()
            self.detected_set_eingang.clear()
            self.bind("<Return>", lambda e: self.on_drucken_eingang())
            self.bind("<Escape>", lambda e: self.show_startseite())

        def build_warenausgang(self):
            tk.Label(self.warenausgang_seite, text="Warenausgang", font=("Arial", 30)).pack(pady=20)
            main_frame = tk.Frame(self.warenausgang_seite)
            main_frame.pack(expand=True, fill="both")

            left_frame = tk.Frame(main_frame)
            left_frame.pack(side="left", expand=True, fill="both", padx=20, pady=20)

            self.dropdown_var_ausgang = tk.StringVar()
            self.dropdown_ausgang = ttk.Combobox(left_frame, textvariable=self.dropdown_var_ausgang, width=30)
            excel_files_ausgang = self.load_excel_files("../ausgang")
            self.dropdown_ausgang['values'] = excel_files_ausgang
            if excel_files_ausgang:
                self.dropdown_ausgang.set(excel_files_ausgang[0])
                filepath = os.path.join("../ausgang", excel_files_ausgang[0])
                self.load_excel_data(filepath, "ausgang")
            self.dropdown_ausgang.bind('<Button-1>', self.refresh_dropdown_ausgang)
            self.dropdown_ausgang.bind('<<ComboboxSelected>>', self.on_excel_select_ausgang)
            self.dropdown_ausgang.pack(pady=10)

            tk.Label(left_frame, text="Erfasste Artikel", font=("Arial", 14)).pack(pady=(20,5))
            columns = ("artikelnummer", "menge", "karton", "beutel", "empfaenger", "status")
            self.tree_ausgang = ttk.Treeview(left_frame, columns=columns, show="headings", height=15)
            self.tree_ausgang.heading("artikelnummer", text="Artikelnummer")
            self.tree_ausgang.heading("menge", text="Menge")
            self.tree_ausgang.heading("karton", text="Karton")
            self.tree_ausgang.heading("beutel", text="Beutel")
            self.tree_ausgang.heading("empfaenger", text="Empfänger")
            self.tree_ausgang.heading("status", text="Status")
            self.tree_ausgang.column("artikelnummer", width=70, anchor="center")
            self.tree_ausgang.column("menge", width=70, anchor="center")
            self.tree_ausgang.column("karton", width=70, anchor="center")
            self.tree_ausgang.column("beutel", width=70, anchor="center")
            self.tree_ausgang.column("empfaenger", width=70, anchor="center")
            self.tree_ausgang.column("status", width=50, anchor="center")
            self.tree_ausgang.pack(expand=True, fill="both")

            self.right_frame_ausgang = tk.Frame(main_frame, bg="black")
            self.right_frame_ausgang.pack(side="right", expand=False, fill="y", padx=20, pady=20)

            button_frame = tk.Frame(self.warenausgang_seite)
            button_frame.pack(pady=10)
            tk.Button(button_frame, text="Drucken", command=self.on_drucken_ausgang).pack(side="left", padx=5)  # NEU
            tk.Button(button_frame, text="Zurück", command=self.show_startseite).pack(side="left", padx=5)
            self.add_logo(self.warenausgang_seite)

        def show_warenausgang(self):
            self.current_page = "ausgang"
            self.startseite.pack_forget()
            self.wareneingang_seite.pack_forget()
            self.warenausgang_seite.pack(expand=True, fill="both")
            self.check_webcam_for_page()
            self.start_webcam_stream(self.right_frame_ausgang)
            self.start_ocr()  # OCR nur hier aktivieren
            self.detected_set_ausgang.clear()  # neue Sitzung für Ausgang
            self.bind("<Return>", lambda e: self.on_drucken_ausgang())  # NEU
            self.bind("<Escape>", lambda e: self.show_startseite())

        #------------------------------------------------------- BASIS FUNKTIONALITÄT ---------------------------------------------------------

        def load_excel_files(self, directory):
            """Lädt alle Excel-Dateien aus dem angegebenen Verzeichnis"""
            excel_files = []
            try:
                if os.path.exists(directory):
                    for file in os.listdir(directory):
                        if file.endswith(('.xlsx', '.xls')):
                            excel_files.append(file)
                else:
                    print(f"Verzeichnis nicht gefunden: {directory}")
            except Exception as e:
                print(f"Fehler beim Laden der Excel-Dateien: {e}")
            return excel_files

        def load_excel_data(self, filepath, page_type="eingang"):
            """Lädt Excel-Datei und speichert Daten in artikel_dict"""
            try:
                workbook = openpyxl.load_workbook(filepath)
                sheet = workbook.active
                
                headers = []
                for cell in sheet[1]:
                    if cell.value:
                        headers.append(cell.value)
                    else:
                        break
                
                data_rows = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if any(row):
                        row_dict = {}
                        for i, value in enumerate(row[:len(headers)]):
                            row_dict[headers[i]] = value if value is not None else ""
                        data_rows.append(row_dict)
                
                if page_type == "eingang":
                    self.artikel_dict_eingang = data_rows
                    self.current_excel_eingang_path = filepath  # ← gemerkt
                    print(f"Wareneingang: {len(data_rows)} Artikel aus Excel geladen")
                else:
                    self.artikel_dict_ausgang = data_rows
                    self.current_excel_ausgang_path = filepath  # ← gemerkt
                    print(f"Warenausgang: {len(data_rows)} Artikel aus Excel geladen")
                
                workbook.close()
                return data_rows
            except Exception as e:
                print(f"Fehler beim Laden der Excel-Datei: {e}")
                return []

        def refresh_dropdown_eingang(self, event=None):
            excel_files = self.load_excel_files("../eingang")
            self.dropdown_eingang['values'] = excel_files

        def refresh_dropdown_ausgang(self, event=None):
            excel_files = self.load_excel_files("../ausgang")
            self.dropdown_ausgang['values'] = excel_files

        def on_excel_select_eingang(self, event=None):
            selected_file = self.dropdown_var_eingang.get()
            if selected_file:
                filepath = os.path.join("../eingang", selected_file)
                self.load_excel_data(filepath, "eingang")

        def on_excel_select_ausgang(self, event=None):
            selected_file = self.dropdown_var_ausgang.get()
            if selected_file:
                filepath = os.path.join("../ausgang", selected_file)
                self.load_excel_data(filepath, "ausgang")

        def find_logitech_c920(self, show_popup=False):
            """Sucht nach der Logitech C920 Webcam über den Gerätenamen"""
            try:
                graph = FilterGraph()
                devices = graph.get_input_devices()
                print(f"Gefundene Geräte: {devices}")
                for device_index, device_name in enumerate(devices):
                    if "c920" in device_name.lower():
                        print(f"Logitech C920 gefunden: {device_name} (Index: {device_index})")
                        return device_index
                if show_popup:
                    self.show_camera_not_found_popup()
                return None
            except Exception as e:
                print(f"Fehler beim Suchen der Webcam: {e}")
                if show_popup:
                    self.show_camera_not_found_popup()
                return None

        def initialize_webcam(self):
            """Initialisiert die C920 mit DirectShow und bestmöglicher Auflösung (max 720p)."""
            camera_index = self.find_logitech_c920(show_popup=False)
            if camera_index is None:
                return False
            try:
                if self.cap is not None:
                    try:
                        self.cap.release()
                    except Exception:
                        pass

                self.cap = cv2.VideoCapture(camera_index, cv2.CAP_DSHOW)
                if not self.cap.isOpened():
                    print("Webcam konnte nicht geöffnet werden")
                    return False

                self.cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc(*'MJPG'))

                selected = None
                for w, h in self.capture_candidates:
                    self.cap.set(cv2.CAP_PROP_FRAME_WIDTH, w)
                    self.cap.set(cv2.CAP_PROP_FRAME_HEIGHT, h)
                    self.cap.set(cv2.CAP_PROP_FPS, self.capture_fps)
                    self.cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
                    self.cap.set(cv2.CAP_PROP_AUTOFOCUS, 1)
                    self.cap.set(cv2.CAP_PROP_AUTO_EXPOSURE, 0.25)

                    rw = int(self.cap.get(cv2.CAP_PROP_FRAME_WIDTH))
                    rh = int(self.cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
                    if rw == w and rh == h:
                        selected = (rw, rh)
                        break

                if selected is None:
                    selected = (int(self.cap.get(cv2.CAP_PROP_FRAME_WIDTH)),
                                int(self.cap.get(cv2.CAP_PROP_FRAME_HEIGHT)))

                print(f"Webcam initialisiert ({selected[0]}x{selected[1]} @ {self.capture_fps} FPS, CAP_DSHOW + MJPG)")
                return True
            except Exception as e:
                print(f"Fehler beim Öffnen der Webcam: {e}")
                return False

        def check_webcam_for_page(self):
            """Prüft ob Webcam verfügbar ist und initialisiert sie bei Bedarf."""
            if self.cap is None or not self.cap.isOpened():
                ok = self.initialize_webcam()
                if not ok:
                    self.show_camera_not_found_popup()

        def start_webcam_stream(self, frame):
            """Startet den Webcam-Livestream im gegebenen Frame"""
            self.webcam_label = tk.Label(frame, text="Webcam wird geladen...",
                                         font=("Arial", 14), bg="lightgray")
            self.webcam_label.pack(expand=True, fill="both")
            self.update_webcam_stream()

        def update_webcam_stream(self):
            """Aktualisiert das Webcam-Bild kontinuierlich"""
            if self.cap is not None and self.cap.isOpened():
                ret, frame = self.cap.read()
                if ret:
                    # Frame für OCR bereitstellen
                    self.last_frame = frame.copy()

                    # OCR-Boxen (rot) zeichnen, falls vorhanden
                    for box in self.ocr_results:
                        x, y, w, h = box['left'], box['top'], box['width'], box['height']
                        cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 255), 2)
                        cv2.putText(frame, box['text'], (x, max(0, y - 5)),
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)

                    frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                    display_frame = cv2.resize(frame_rgb, (640, 480))
                    img = Image.fromarray(display_frame)
                    photo = ImageTk.PhotoImage(img)
                    if hasattr(self, 'webcam_label'):
                        self.webcam_label.configure(image=photo, text="")
                        self.webcam_label.image = photo
                else:
                    if hasattr(self, 'webcam_label'):
                        self.webcam_label.configure(text="Webcam-Fehler", image="")
            else:
                if hasattr(self, 'webcam_label'):
                    self.webcam_label.configure(text="Keine Webcam gefunden", image="")

            if hasattr(self, 'webcam_label'):
                self.after(33, self.update_webcam_stream)

        def stop_webcam_stream(self):
            """Stoppt den Webcam-Stream"""
            if hasattr(self, 'webcam_label'):
                self.webcam_label.destroy()
                delattr(self, 'webcam_label')

        def show_camera_not_found_popup(self):
            """Zeigt ein Pop-up an, wenn die Logitech C920 nicht gefunden wurde"""
            popup = tk.Toplevel(self)
            popup.title("Kamera nicht gefunden")
            popup.geometry("300x150")
            popup.resizable(False, False)
            popup.transient(self)
            popup.grab_set()
            message_label = tk.Label(popup, text="Logitech C920 nicht gefunden",
                                   font=("Arial", 12), pady=20)
            message_label.pack()
            ok_button = tk.Button(popup, text="OK", command=popup.destroy,
                                width=10, pady=5)
            ok_button.pack(pady=10)
            popup.focus_set()

        # --- OCR-Lebenszyklus ---
        def start_ocr(self):
            """Startet die einfache Vollbild-OCR nur auf Eingangs-/Ausgangsseiten."""
            if self.current_page not in ("eingang", "ausgang"):
                return
            if self.stop_event and not self.stop_event.is_set():
                return  # schon aktiv
            self.stop_event = threading.Event()
            threading.Thread(target=self.ocr_loop, daemon=True).start()

        def stop_ocr(self):
            """Stoppt die OCR (falls aktiv)."""
            if self.stop_event:
                try:
                    self.stop_event.set()
                except Exception:
                    pass
                self.stop_event = None
            self.ocr_results = []
            # Set leeren, damit beim nächsten Aufruf neu erkannt werden kann
            if hasattr(self, "detected_set_eingang"):
                self.detected_set_eingang.clear()

        def ocr_loop(self):
            """Rotationstolerante Vollbild-OCR (A-Z, 0-9) mit Bounding-Boxes."""
            base_config_cardinal = '--oem 3 --psm 6 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789'
            base_config_sweep = '--oem 3 --psm 11 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789'
            conf_threshold = 50
            conf_threshold_sweep = 48
            angles_cardinal = [0, 180, 90, 270]
            delta = [-12, -8, -4, 4, 8, 12]
            angles_sweep = [a + d for a in angles_cardinal for d in delta]
            early_stop_boxes = 10
            target_period = 0.25  # ← max. 4 OCR-Durchläufe pro Sekunde

            def rotate_with_bounds(img, angle):
                (h, w) = img.shape[:2]
                (cx, cy) = (w / 2.0, h / 2.0)
                M = cv2.getRotationMatrix2D((cx, cy), angle, 1.0)
                cos = abs(M[0, 0]); sin = abs(M[0, 1])
                nW = int(h * sin + w * cos)
                nH = int(h * cos + w * sin)
                M[0, 2] += (nW / 2.0) - cx
                M[1, 2] += (nH / 2.0) - cy
                rotated = cv2.warpAffine(img, M, (nW, nH), flags=cv2.INTER_LINEAR, borderMode=cv2.BORDER_REPLICATE)
                Minv = cv2.invertAffineTransform(M)
                return rotated, Minv

            def map_box_back(Minv, x, y, w, h, sw, sh):
                pts = [(x, y), (x + w, y), (x, y + h), (x + w, y + h)]
                mapped = []
                for px, py in pts:
                    ox = Minv[0, 0] * px + Minv[0, 1] * py + Minv[0, 2]
                    oy = Minv[1, 0] * px + Minv[1, 1] * py + Minv[1, 2]
                    mapped.append((ox, oy))
                xs = [p[0] for p in mapped]; ys = [p[1] for p in mapped]
                x_min = max(0, int(np.floor(min(xs))))
                y_min = max(0, int(np.floor(min(ys))))
                x_max = min(sw, int(np.ceil(max(xs))))
                y_max = min(sh, int(np.ceil(max(ys))))
                W = max(0, x_max - x_min); H = max(0, y_max - y_min)
                return x_min, y_min, W, H

            while self.stop_event and not self.stop_event.is_set():
                loop_start = time.perf_counter()  # ← Startzeit für Throttling
                frame = self.last_frame
                if frame is None or self.current_page not in ("eingang", "ausgang"):
                    # Bei inaktivem Frame/Page trotzdem nicht über 4 Hz laufen
                    self.stop_event.wait(target_period)
                    continue
                try:
                    orig_h, orig_w = frame.shape[:2]
                    target_w = min(self.ocr_max_width, orig_w)
                    scale = target_w / float(orig_w)
                    small = cv2.resize(frame, (target_w, int(orig_h * scale)))
                    small_h, small_w = small.shape[:2]

                    boxes_all = []
                    texts_all = []

                    # 1) Kardinale Winkel (PSM 6, ohne Binarisierung)
                    for angle in angles_cardinal:
                        rot, Minv = rotate_with_bounds(small, angle)
                        rot_rgb = cv2.cvtColor(rot, cv2.COLOR_BGR2RGB)
                        data = pytesseract.image_to_data(rot_rgb, config=base_config_cardinal, output_type=pytesseract.Output.DICT)
                        n = len(data.get('text', []))
                        for i in range(n):
                            text = (data['text'][i] or '').strip().upper()
                            if not text or not re.fullmatch(r'[A-Z0-9]+', text):
                                continue
                            try:
                                conf = int(float(data['conf'][i]))
                            except ValueError:
                                conf = -1
                            if conf < conf_threshold:
                                continue
                            rx = int(data['left'][i]); ry = int(data['top'][i])
                            rw = int(data['width'][i]); rh = int(data['height'][i])
                            sx, sy, sw_box, sh_box = map_box_back(Minv, rx, ry, rw, rh, small_w, small_h)
                            if sw_box <= 2 or sh_box <= 2:
                                continue
                            ox = int(round(sx / scale)); oy = int(round(sy / scale))
                            ow = int(round(sw_box / scale)); oh = int(round(sh_box / scale))
                            ox = max(0, min(ox, orig_w - 1)); oy = max(0, min(oy, orig_h - 1))
                            if ox + ow > orig_w: ow = orig_w - ox
                            if oy + oh > orig_h: oh = orig_h - oy
                            if ow <= 2 or oh <= 2:
                                continue
                            boxes_all.append({'text': text, 'left': ox, 'top': oy, 'width': ow, 'height': oh})
                            texts_all.append(text)

                    # 2) Feinsweeps (PSM 11 + Otsu), nur wenn noch wenig Treffer
                    if len(boxes_all) < early_stop_boxes:
                        for angle in angles_sweep:
                            if len(boxes_all) >= early_stop_boxes:
                                break
                            rot, Minv = rotate_with_bounds(small, angle)
                            gray = cv2.cvtColor(rot, cv2.COLOR_BGR2GRAY)
                            _, bin_img = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                            data = pytesseract.image_to_data(bin_img, config=base_config_sweep, output_type=pytesseract.Output.DICT)
                            n = len(data.get('text', []))
                            for i in range(n):
                                text = (data['text'][i] or '').strip().upper()
                                if not text or not re.fullmatch(r'[A-Z0-9]+', text):
                                    continue
                                try:
                                    conf = int(float(data['conf'][i]))
                                except ValueError:
                                    conf = -1
                                if conf < conf_threshold_sweep:
                                    continue
                                rx = int(data['left'][i]); ry = int(data['top'][i])
                                rw = int(data['width'][i]); rh = int(data['height'][i])
                                sx, sy, sw_box, sh_box = map_box_back(Minv, rx, ry, rw, rh, small_w, small_h)
                                if sw_box <= 2 or sh_box <= 2:
                                    continue
                                ox = int(round(sx / scale)); oy = int(round(sy / scale))
                                ow = int(round(sw_box / scale)); oh = int(round(sh_box / scale))
                                ox = max(0, min(ox, orig_w - 1)); oy = max(0, min(oy, orig_h - 1))
                                if ox + ow > orig_w: ow = orig_w - ox
                                if oy + oh > orig_h: oh = orig_h - oy
                                if ow <= 2 or oh <= 2:
                                    continue
                                boxes_all.append({'text': text, 'left': ox, 'top': oy, 'width': ow, 'height': oh})
                                texts_all.append(text)

                    # Duplikate reduzieren
                    dedup = []
                    seen = set()
                    for b in boxes_all:
                        key = (b['text'], b['left'] // 25, b['top'] // 25)
                        if key in seen:
                            continue
                        seen.add(key)
                        dedup.append(b)

                    self.ocr_results = dedup

                    # Wareneingang: erkannte Texte mit Excel (Status "0") abgleichen und einfügen
                    if self.current_page == "eingang":
                        for t in sorted(set(texts_all)):
                            row_match = self.find_eingang_match(t)
                            if row_match:
                                self.after(0, self.insert_eingang_row, row_match)

                    # Warenausgang: erkannte Texte mit Excel (Status "0") abgleichen und einfügen (inkl. Empfänger)
                    if self.current_page == "ausgang":
                        for t in sorted(set(texts_all)):
                            row_match = self.find_ausgang_match(t)
                            if row_match:
                                self.after(0, self.insert_ausgang_row, row_match)

                    if texts_all:
                        print("Erkannt:", ", ".join(sorted(set(texts_all))))
                except Exception as e:
                    print(f"OCR-Fehler: {e}")
                # Dynamische Wartezeit, sodass ein Loop mind. target_period dauert
                elapsed = time.perf_counter() - loop_start
                remaining = target_period - elapsed
                if remaining > 0:
                    self.stop_event.wait(remaining)

        # ------------------------ OCR → Wareneingang: Treffer verarbeiten ------------------------
        def _norm_text(self, s: str) -> str:
            """Normalisiert Artikelnummern: Großbuchstaben, ohne Leer-/Bindestriche."""
            return (s or "").replace(" ", "").replace("-", "").upper()

        def _row_get(self, row: dict, *keys):
            """Liefert den ersten vorhandenen Wert aus row für die gegebenen Keys (case-insensitiv)."""
            if not row:
                return ""
            lower_map = {str(k).lower(): v for k, v in row.items()}
            for k in keys:
                v = lower_map.get(str(k).lower())
                if v not in (None, ""):
                    return v
            return ""

        def find_eingang_match(self, ocr_text: str):
            """Sucht in artikel_dict_eingang nach Status '0' und passender Artikelnummer."""
            if not self.artikel_dict_eingang:
                return None
            ocr_norm = self._norm_text(ocr_text)
            for row in self.artikel_dict_eingang:
                status = str(self._row_get(row, "Status")).strip()
                if status != "0":
                    continue
                art = self._row_get(row, "Artikelnummer")
                if self._norm_text(str(art)) == ocr_norm:
                    return row
            return None

        def insert_eingang_row(self, row: dict):
            """Fügt einen Wareneingang-Datensatz in die Tabelle ein (falls noch nicht vorhanden) und aktualisiert Karton/Beutel aus DB."""
            art = str(self._row_get(row, "Artikelnummer"))
            art_norm = self._norm_text(art)
            if art_norm in self.detected_set_eingang:
                return
            menge  = str(self._row_get(row, "Menge"))
            karton = str(self._row_get(row, "Karton"))
            beutel = str(self._row_get(row, "Beutel"))
            status = str(self._row_get(row, "Status"))
            try:
                iid = self.tree_eingang.insert("", "end", values=(art, menge, karton, beutel, status))
                self.detected_set_eingang.add(art_norm)
                # Karton/Beutel aus Datenbank aktualisieren (asynchron, UI bleibt flüssig)
                self.update_row_from_db_async(art_norm, iid)
            except Exception as e:
                print(f"[WARN] Konnte Datensatz nicht einfügen: {e}")

        def update_row_from_db_async(self, art_norm: str, iid: str):
            """(Bestehend) Variante für Wareneingang: ruft die generische Methode auf."""
            # Cache-Hit
            if art_norm in self.db_cache_karton_beutel:
                k, b = self.db_cache_karton_beutel[art_norm]
                self._apply_db_values_to_treeview(self.tree_eingang, iid, k, b)
                return
            def worker():
                k, b = self.query_db_karton_beutel(art_norm)
                if k is not None or b is not None:
                    self.db_cache_karton_beutel[art_norm] = (k, b)
                    self.after(0, self._apply_db_values_to_treeview, self.tree_eingang, iid, k, b)
            threading.Thread(target=worker, daemon=True).start()

        def query_db_karton_beutel(self, art_norm: str):
            """Liest Karton/Beutel aus der Access-DB [Artikeldatenbank] für die gegebene normierte Artikelnummer."""
            try:
                if not os.path.exists(self.db_path):
                    print(f"[DB] Datei nicht gefunden: {self.db_path}")
                    return (None, None)
                conn = pyodbc.connect(f"Driver={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={self.db_path};")
                try:
                    cur = conn.cursor()
                    # Vergleich auf normierter Artikelnummer (ohne Leer-/Bindestriche, uppercase)
                    sql = """
                        SELECT [Karton], [Beutel]
                        FROM [Artikeldatenbank]
                        WHERE UCASE(REPLACE(REPLACE([Artikelnummer], '-', ''), ' ', '')) = ?
                    """
                    cur.execute(sql, art_norm)
                    row = cur.fetchone()
                    if row:
                        k = "" if row[0] is None else str(row[0])
                        b = "" if row[1] is None else str(row[1])
                        return (k, b)
                    return (None, None)
                finally:
                    conn.close()
            except Exception as e:
                print(f"[DB] Fehler beim Zugriff: {e}")
                return (None, None)

        def _apply_db_values_to_tree(self, iid: str, karton: str, beutel: str):
            """Trägt Karton/Beutel aus DB in die bestehende Zeile ein (überschreibt nur diese Felder)."""
            try:
                vals = list(self.tree_eingang.item(iid, "values"))
                if not vals or len(vals) < 5:
                    return
                # Indizes: 0=Artikelnummer, 1=Menge, 2=Karton, 3=Beutel, 4=Status
                if karton is not None and karton != "":
                    vals[2] = karton
                if beutel is not None and beutel != "":
                    vals[3] = beutel
                self.tree_eingang.item(iid, values=tuple(vals))
            except Exception as e:
                print(f"[DB] Fehler beim Aktualisieren der Zeile: {e}")

        def query_db_karton_beutel(self, art_norm: str):
            """Liest Karton/Beutel aus der Access-DB [Artikeldatenbank] für die gegebene normierte Artikelnummer."""
            try:
                if not os.path.exists(self.db_path):
                    print(f"[DB] Datei nicht gefunden: {self.db_path}")
                    return (None, None)

                # passenden Access-Treiber wählen
                drivers = [d for d in pyodbc.drivers() if "Access Driver" in d]
                if not drivers:
                    print("[DB] Access-ODBC-Treiber fehlt. Installiere 'Microsoft Access Database Engine (64-bit)'.")
                    return (None, None)
                driver = drivers[-1]  # jüngsten nehmen

                conn = pyodbc.connect(f"DRIVER={{{driver}}};DBQ={self.db_path};")
                try:
                    cur = conn.cursor()
                    sql = """
                        SELECT [Karton], [Beutel]
                        FROM [Artikeldatenbank]
                        WHERE UCASE(REPLACE(REPLACE([Artikelnummer], '-', ''), ' ', '')) = ?
                    """
                    cur.execute(sql, art_norm)
                    row = cur.fetchone()
                    if row:
                        k = "" if row[0] is None else str(row[0])
                        b = "" if row[1] is None else str(row[1])
                        return (k, b)
                    return (None, None)
                finally:
                    conn.close()
            except Exception as e:
                print(f"[DB] Fehler beim Zugriff: {e}")
                return (None, None)

        # ------------------------ OCR → Warenausgang: Treffer verarbeiten ------------------------
        def find_ausgang_match(self, ocr_text: str):
            """Sucht in artikel_dict_ausgang nach Status '0' und passender Artikelnummer."""
            if not self.artikel_dict_ausgang:
                return None
            ocr_norm = self._norm_text(ocr_text)
            for row in self.artikel_dict_ausgang:
                status = str(self._row_get(row, "Status")).strip()
                if status != "0":
                    continue
                art = self._row_get(row, "Artikelnummer")
                if self._norm_text(str(art)) == ocr_norm:
                    return row
            return None

        def insert_ausgang_row(self, row: dict):
            """Fügt einen Warenausgang-Datensatz in die Tabelle ein (falls noch nicht vorhanden) und aktualisiert Karton/Beutel aus DB."""
            art = str(self._row_get(row, "Artikelnummer"))
            art_norm = self._norm_text(art)
            if art_norm in self.detected_set_ausgang:
                return
            menge      = str(self._row_get(row, "Menge"))
            karton     = str(self._row_get(row, "Karton"))
            beutel     = str(self._row_get(row, "Beutel"))
            empfaenger = str(self._row_get(row, "Empfänger"))
            status     = str(self._row_get(row, "Status"))
            try:
                iid = self.tree_ausgang.insert("", "end", values=(art, menge, karton, beutel, empfaenger, status))
                self.detected_set_ausgang.add(art_norm)
                # Karton/Beutel aus Datenbank aktualisieren (asynchron)
                self.update_row_from_db_async_for_tree(art_norm, self.tree_ausgang, iid)
            except Exception as e:
                print(f"[WARN] Konnte Ausgangs-Datensatz nicht einfügen: {e}")

        def update_row_from_db_async_for_tree(self, art_norm: str, tree: ttk.Treeview, iid: str):
            """Wie update_row_from_db_async, nur generisch für beliebigen Treeview (Eingang/Ausgang)."""
            if art_norm in self.db_cache_karton_beutel:
                k, b = self.db_cache_karton_beutel[art_norm]
                self._apply_db_values_to_treeview(tree, iid, k, b)
                return
            def worker():
                k, b = self.query_db_karton_beutel(art_norm)
                if k is not None or b is not None:
                    self.db_cache_karton_beutel[art_norm] = (k, b)
                    self.after(0, self._apply_db_values_to_treeview, tree, iid, k, b)
            threading.Thread(target=worker, daemon=True).start()

        def _apply_db_values_to_treeview(self, tree: ttk.Treeview, iid: str, karton: str, beutel: str):
            """Trägt Karton/Beutel in eine bestehende Zeile eines Treeviews ein (Eingang/Ausgang)."""
            try:
                vals = list(tree.item(iid, "values"))
                if not vals:
                    return
                # Eingangs-Tabelle: [Artikelnr, Menge, Karton, Beutel, Status]
                # Ausgangs-Tabelle: [Artikelnr, Menge, Karton, Beutel, Empfänger, Status]
                # Karton = Index 2, Beutel = Index 3 bei beiden Tabellen
                if len(vals) >= 4:
                    if karton is not None and karton != "":
                        vals[2] = karton
                    if beutel is not None and beutel != "":
                        vals[3] = beutel
                    tree.item(iid, values=tuple(vals))
            except Exception as e:
                print(f"[DB] Fehler beim Aktualisieren der Zeile (generic): {e}")

        # ---------------------------------------- Drucken-Funktionen ----------------------------------------

        def on_drucken_eingang(self):
            """Druck-Workflow Wareneingang:
            - Artikelnummer/Menge in Konsole ausgeben
            - Status in Excel (0→1) setzen
            - In Tabelle Status 0→✅; war bereits ✅, Zeile entfernen
            """
            # Auswahl holen; wenn keine Auswahl, alle Zeilen verarbeiten
            iids = self.tree_eingang.selection()
            if not iids:
                iids = self.tree_eingang.get_children()

            for iid in list(iids):
                vals = list(self.tree_eingang.item(iid, "values"))
                if not vals or len(vals) < 5:
                    continue
                artikelnummer = str(vals[0])
                menge = str(vals[1])
                status_val = str(vals[4])

                print(f"Artikelnummer: {artikelnummer} | Menge: {menge}")

                # War Status bereits ein grüner Haken? → Zeile entfernen
                if status_val in ("✅", "✔️", "1"):
                    try:
                        self.tree_eingang.delete(iid)
                    except Exception:
                        pass
                    continue

                # Status == 0 → Excel aktualisieren und UI auf ✅ setzen
                if status_val.strip() == "0":
                    updated_excel = self._excel_set_status_eingang(artikelnummer, "1")
                    # UI aktualisieren (unabhängig vom Excel-Ergebnis setzen wir den Haken)
                    vals[4] = "✅"
                    self.tree_eingang.item(iid, values=tuple(vals))
                    # internen Cache anpassen
                    self._update_internal_eingang_status(artikelnummer, "1")

        def on_drucken_ausgang(self):
            """Druck-Workflow Warenausgang:
            - Artikelnummer/Menge in Konsole ausgeben
            - Status in Excel (0→1) setzen
            - In Tabelle Status 0→✅; war bereits ✅, Zeile entfernen
            """
            iids = self.tree_ausgang.selection()
            if not iids:
                iids = self.tree_ausgang.get_children()

            for iid in list(iids):
                vals = list(self.tree_ausgang.item(iid, "values"))
                if not vals or len(vals) < 6:
                    continue
                artikelnummer = str(vals[0])
                menge = str(vals[1])
                status_val = str(vals[5])

                print(f"Artikelnummer: {artikelnummer} | Menge: {menge}")

                # Bereits erledigt? → Zeile entfernen
                if status_val in ("✅", "✔️", "1"):
                    try:
                        self.tree_ausgang.delete(iid)
                    except Exception:
                        pass
                    continue

                # Status == 0 → Excel aktualisieren und UI auf ✅ setzen
                if status_val.strip() == "0":
                    updated_excel = self._excel_set_status_ausgang(artikelnummer, "1")
                    vals[5] = "✅"
                    self.tree_ausgang.item(iid, values=tuple(vals))
                    self._update_internal_ausgang_status(artikelnummer, "1")

        def _excel_set_status_eingang(self, artikelnummer: str, new_status: str) -> bool:
            """Setzt den Status (z.B. 1) in der Eingangs-Excel für die gegebene Artikelnummer.
            Gibt True zurück, wenn in der Datei aktualisiert wurde.
            """
            path = getattr(self, "current_excel_eingang_path", None)
            if not path or not os.path.exists(path):
                print("[Excel] Keine Eingangs-Excel gesetzt oder Datei fehlt.")
                return False
            try:
                wb = openpyxl.load_workbook(path)
                ws = wb.active
                # Header-Mapping
                headers = {}
                for idx, cell in enumerate(ws[1], start=1):
                    if cell.value:
                        headers[str(cell.value).strip().lower()] = idx
                col_art = headers.get("artikelnummer")
                col_status = headers.get("status")
                if not col_art or not col_status:
                    print("[Excel] Spalten 'Artikelnummer' oder 'Status' nicht gefunden.")
                    wb.close()
                    return False

                art_norm_target = self._norm_text(artikelnummer)
                updated = False
                for r in range(2, ws.max_row + 1):
                    art_val = ws.cell(row=r, column=col_art).value
                    if self._norm_text(str(art_val)) != art_norm_target:
                        continue
                    status_cell = ws.cell(row=r, column=col_status)
                    cur = "" if status_cell.value is None else str(status_cell.value).strip()
                    if cur == "0":
                        status_cell.value = new_status
                        updated = True
                        break
                if updated:
                    wb.save(path)
                wb.close()
                return updated
            except Exception as e:
                print(f"[Excel] Fehler beim Aktualisieren: {e}")
                return False

        def _excel_set_status_ausgang(self, artikelnummer: str, new_status: str) -> bool:
            """Setzt den Status (z.B. 1) in der Ausgangs-Excel für die gegebene Artikelnummer."""
            path = getattr(self, "current_excel_ausgang_path", None)
            if not path or not os.path.exists(path):
                print("[Excel] Keine Ausgangs-Excel gesetzt oder Datei fehlt.")
                return False
            try:
                wb = openpyxl.load_workbook(path)
                ws = wb.active
                # Header-Mapping
                headers = {}
                for idx, cell in enumerate(ws[1], start=1):
                    if cell.value:
                        headers[str(cell.value).strip().lower()] = idx
                col_art = headers.get("artikelnummer")
                col_status = headers.get("status")
                if not col_art or not col_status:
                    print("[Excel] Spalten 'Artikelnummer' oder 'Status' nicht gefunden (Ausgang).")
                    wb.close()
                    return False

                art_norm_target = self._norm_text(artikelnummer)
                updated = False
                for r in range(2, ws.max_row + 1):
                    art_val = ws.cell(row=r, column=col_art).value
                    if self._norm_text(str(art_val)) != art_norm_target:
                        continue
                    status_cell = ws.cell(row=r, column=col_status)
                    cur = "" if status_cell.value is None else str(status_cell.value).strip()
                    if cur == "0":
                        status_cell.value = new_status
                        updated = True
                        break
                if updated:
                    wb.save(path)
                wb.close()
                return updated
            except Exception as e:
                print(f"[Excel] Fehler beim Aktualisieren (Ausgang): {e}")
                return False

        def _update_internal_eingang_status(self, artikelnummer: str, new_status: str):
            """Spiegelt den geänderten Status in artikel_dict_eingang wider."""
            target = self._norm_text(artikelnummer)
            for row in self.artikel_dict_eingang:
                art = self._row_get(row, "Artikelnummer")
                if self._norm_text(str(art)) == target:
                    row["Status"] = new_status
                    break

        def _update_internal_ausgang_status(self, artikelnummer: str, new_status: str):
            """Spiegelt den geänderten Status in artikel_dict_ausgang wider."""
            target = self._norm_text(artikelnummer)
            for row in self.artikel_dict_ausgang:
                art = self._row_get(row, "Artikelnummer")
                if self._norm_text(str(art)) == target:
                    row["Status"] = new_status
                    break

        # ---------------------------------------- Platzhalter-Funktionen ----------------------------------------

        def find_printer(self):
            """Placeholder für Find Printer-Funktionalität"""
            print("Find Printer-Funktion aufgerufen")

        # ---------------------------------------- Debugging-Hilfen ----------------------------------------
        def debug_print_article_dicts(self):
            print(f"[DEBUG] Eingang: {len(self.artikel_dict_eingang)} Einträge")
            for i, row in enumerate(self.artikel_dict_eingang[:5]):
                print(f"[E{i}] {row}")
            print(f"[DEBUG] Ausgang: {len(self.artikel_dict_ausgang)} Einträge")
            for i, row in enumerate(self.artikel_dict_ausgang[:5]):
                print(f"[A{i}] {row}")

if __name__ == "__main__":
    print("=== WEBCAM-ANSICHT MIT EINFACHER VOLLBILD-OCR (A-Z,0-9) ===")
    app = App()
    app.mainloop()